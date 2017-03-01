from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'hour'
ws['B1'] = 'minute'
ws['C1'] = 'second'

# Rows can also be appended
ws.append([1, 2, 3])

# Save the file
wb.save("time.xlsx")
