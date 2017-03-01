from openpyxl import Workbook
import time
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'hour'
ws['B1'] = 'minute'
ws['C1'] = 'second'

while 1:
    old_time = time.strftime("%H:%M:%S")
    while 1:
        new_time = time.strftime("%H:%M:%S")
        if new_time != old_time:
            old_time = new_time
            ws.append(old_time.split(':'))
            wb.save("time.xlsx")
            print(old_time)
            break

